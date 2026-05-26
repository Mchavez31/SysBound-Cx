"""
Compare two drawing PDFs and produce structured results + Excel export.
"""
from __future__ import annotations

import json
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook

ProgressCb = Callable[[int, str], None] | None
from openpyxl.styles import PatternFill

from services import pdf_parser


def _view_fields_for_tag(tag: dict[str, Any] | None, suffix: str) -> dict[str, Any]:
    """Tag hit box center + page size for focused PDF/snippet viewers (pdfplumber coordinates)."""
    if not tag:
        return {}
    out: dict[str, Any] = {}
    for src, dest in (
        ("x", f"tag_x_{suffix}"),
        ("y", f"tag_y_{suffix}"),
        ("page_width", f"page_width_{suffix}"),
        ("page_height", f"page_height_{suffix}"),
    ):
        v = tag.get(src)
        if v is None:
            continue
        try:
            out[dest] = float(v)
        except (TypeError, ValueError):
            pass
    return out


def _norm_hex(h: str | None) -> str | None:
    if not h:
        return None
    return h.strip().upper()


def _norm_sub(s: str | None) -> str | None:
    if s is None:
        return None
    t = str(s).strip()
    return t if t else None


def _action_for_row(
    change_type: str,
    is_x: bool,
    *,
    has_future: bool = False,
) -> str:
    if is_x:
        return "x_tag"
    if has_future:
        return "future"
    if change_type == "new":
        return "assign_subsystem"
    if change_type == "removed":
        return "review_removal"
    if change_type in ("subsystem_changed", "color_changed"):
        return "review_change"
    return "none"


def compare_systemized_vs_systemized(
    path_a: str,
    path_b: str,
    drawing_a_id: str,
    drawing_b_id: str,
    project_id: str,
    drawing_number_a: str,
    drawing_number_b: str,
    progress_cb: ProgressCb = None,
    tag_exclude_normalized: frozenset[str] | None = None,
) -> dict[str, Any]:
    """
    Compare two systemized PDFs by subsystem assignment per tag (not raw tag equality).
    """
    # Load color palette for this project to map colors to subsystems
    from database_connection import SessionLocal
    from models.database import ColorPalette
    
    color_to_subsystem = {}
    try:
        with SessionLocal() as db:
            palettes = db.query(ColorPalette).filter(ColorPalette.project_id == project_id).all()
            for p in palettes:
                if p.hex_color:
                    color_to_subsystem[p.hex_color.upper()] = p.subsystem_number
            import logging
            logging.warning(f"Loaded {len(color_to_subsystem)} color-to-subsystem mappings from palette")
    except Exception as e:
        import logging
        logging.warning(f"Could not load color palette: {e}")
    
    def _hex_to_rgb(hex_color: str) -> tuple[int, int, int] | None:
        """Convert hex color to RGB tuple."""
        if not hex_color:
            return None
        h = hex_color.lstrip('#').upper()
        if len(h) != 6:
            return None
        try:
            return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
        except:
            return None
    
    def _color_distance(rgb1: tuple[int, int, int], rgb2: tuple[int, int, int]) -> float:
        """Calculate Euclidean distance between two RGB colors."""
        return ((rgb1[0] - rgb2[0]) ** 2 + (rgb1[1] - rgb2[1]) ** 2 + (rgb1[2] - rgb2[2]) ** 2) ** 0.5
    
    def _subsystem_from_color(hex_color: str | None, tolerance: int = 30) -> tuple[str | None, float | None, str | None]:
        """
        Look up subsystem from color using the project's color palette with fuzzy matching.
        Returns: (subsystem, distance, matched_color)
        """
        if not hex_color or not color_to_subsystem:
            return (None, None, None)
        
        # Try exact match first
        normalized = hex_color.upper().lstrip('#')
        if not normalized.startswith('#'):
            normalized = '#' + normalized
        if normalized in color_to_subsystem:
            return (color_to_subsystem[normalized], 0.0, normalized)
        
        # Fuzzy match - find closest color within tolerance
        tag_rgb = _hex_to_rgb(hex_color)
        if not tag_rgb:
            return (None, None, None)
        
        best_match = None
        best_distance = float('inf')
        best_color = None
        
        for palette_color, subsystem in color_to_subsystem.items():
            palette_rgb = _hex_to_rgb(palette_color)
            if palette_rgb:
                dist = _color_distance(tag_rgb, palette_rgb)
                if dist < best_distance and dist <= tolerance:
                    best_distance = dist
                    best_match = subsystem
                    best_color = palette_color
        
        return (best_match, best_distance if best_match else None, best_color)
    
    if progress_cb:
        progress_cb(2, "Reading drawings (parallel)…")

    completion = {"a": 0.0, "b": 0.0}
    prog_lock = threading.Lock()

    def _wrap_progress(side: str, lo: int, hi: int):
        span = max(1, hi - lo)

        def inner(p: int, msg: str) -> None:
            clamped = min(max(int(p), lo), hi)
            frac = (clamped - lo) / span
            with prog_lock:
                completion[side] = max(completion[side], min(1.0, frac))
                # Weighted progress: extraction is 85% of work, comparison is 15%
                overall = 3 + int(82 * (completion["a"] + completion["b"]) / 2)
                overall = min(85, max(3, overall))
                if progress_cb:
                    progress_cb(overall, msg)

        return inner

    with ThreadPoolExecutor(max_workers=2) as pool:
        fut_a = pool.submit(
            pdf_parser.extract_all,
            path_a,
            progress_cb=_wrap_progress("a", 3, 45),
            progress_lo=3,
            progress_hi=45,
            progress_label="Drawing A",
            tag_exclude_normalized=tag_exclude_normalized,
        )
        fut_b = pool.submit(
            pdf_parser.extract_all,
            path_b,
            progress_cb=_wrap_progress("b", 45, 85),
            progress_lo=45,
            progress_hi=85,
            progress_label="Drawing B",
            tag_exclude_normalized=tag_exclude_normalized,
        )
        result_a = fut_a.result()
        result_b = fut_b.result()

    tags_a = {t["tag_number"]: t for t in result_a["tags"]}
    tags_b = {t["tag_number"]: t for t in result_b["tags"]}

    all_nums = set(tags_a.keys()) | set(tags_b.keys())
    only_a = set(tags_a.keys()) - set(tags_b.keys())
    only_b = set(tags_b.keys()) - set(tags_a.keys())
    both = set(tags_a.keys()) & set(tags_b.keys())

    rows: list[dict[str, Any]] = []
    unchanged_count = 0

    sorted_nums = sorted(all_nums)
    n_tag = len(sorted_nums)
    for i, tn in enumerate(sorted_nums):
        if progress_cb and n_tag > 0 and (i == 0 or i == n_tag - 1 or i % max(1, n_tag // 50) == 0):
            p = 86 + int(12 * i / max(n_tag, 1))
            progress_cb(min(98, p), f"Comparing tags ({i + 1}/{n_tag})…")
        in_a = tn in tags_a
        in_b = tn in tags_b
        ta = tags_a.get(tn)
        tb = tags_b.get(tn)

        if in_a and not in_b:
            is_x = bool(ta and ta.get("is_x_tag"))
            rows.append(
                {
                    "tag_number": tn,
                    "tag_type": (ta or {}).get("tag_type") or "other",
                    "change_type": "removed",
                    "action_needed": "review_removal",
                    "subsystem_a": _norm_sub((ta or {}).get("nearest_subsystem")),
                    "subsystem_b": None,
                    "color_a": _norm_hex((ta or {}).get("nearest_subsystem_color")),
                    "color_b": None,
                    "drawing_side": "A",
                    "page_a": (ta or {}).get("page_number"),
                    "page_b": None,
                    "is_x_tag": is_x,
                    "drawing_id_a": drawing_a_id,
                    "drawing_id_b": drawing_b_id,
                    "drawing_number_a": drawing_number_a or "",
                    "drawing_number_b": drawing_number_b or "",
                    **_view_fields_for_tag(ta, "a"),
                }
            )
            continue

        if in_b and not in_a:
            is_x = bool(tb and tb.get("is_x_tag"))
            rows.append(
                {
                    "tag_number": tn,
                    "tag_type": (tb or {}).get("tag_type") or "other",
                    "change_type": "new",
                    "action_needed": "x_tag" if is_x else "assign_subsystem",
                    "subsystem_a": None,
                    "subsystem_b": _norm_sub((tb or {}).get("nearest_subsystem")),
                    "color_a": None,
                    "color_b": _norm_hex((tb or {}).get("nearest_subsystem_color")),
                    "drawing_side": "B",
                    "page_a": None,
                    "page_b": (tb or {}).get("page_number"),
                    "is_x_tag": is_x,
                    "drawing_id_a": drawing_a_id,
                    "drawing_id_b": drawing_b_id,
                    "drawing_number_a": drawing_number_a or "",
                    "drawing_number_b": drawing_number_b or "",
                    **_view_fields_for_tag(tb, "b"),
                }
            )
            continue

        assert ta is not None and tb is not None
        sub_a = _norm_sub(ta.get("nearest_subsystem"))
        sub_b = _norm_sub(tb.get("nearest_subsystem"))
        col_a = _norm_hex(ta.get("nearest_subsystem_color"))
        col_b = _norm_hex(tb.get("nearest_subsystem_color"))
        
        # ALSO compare actual tag bubble fill colors (not just nearest label colors)
        # This catches cases where tag bubble color changed but tag didn't move near a different label
        tag_color_a = _norm_hex(ta.get("tag_fill_color"))
        tag_color_b = _norm_hex(tb.get("tag_fill_color"))
        
        # NEW: Use color palette to determine subsystem from actual tag color
        # This is more reliable than spatial label matching
        # Returns: (subsystem, distance, matched_palette_color)
        sub_from_color_a_result = _subsystem_from_color(tag_color_a)
        sub_from_color_b_result = _subsystem_from_color(tag_color_b)
        
        sub_from_color_a = sub_from_color_a_result[0]
        sub_from_color_b = sub_from_color_b_result[0]
        color_distance_a = sub_from_color_a_result[1]
        color_distance_b = sub_from_color_b_result[1]
        matched_palette_color_a = sub_from_color_a_result[2]
        matched_palette_color_b = sub_from_color_b_result[2]
        
        # Prefer palette-based subsystem if available, otherwise fall back to label-based
        final_sub_a = sub_from_color_a if sub_from_color_a else sub_a
        final_sub_b = sub_from_color_b if sub_from_color_b else sub_b
        
        is_x = bool(ta.get("is_x_tag") or tb.get("is_x_tag"))

        # Debug logging for color detection
        if tn in ["S-98840", "S-98841", "S-98842", "V-98815", "XV-9881521", "XV-9881523"]:
            import logging
            dist_a_str = f"{color_distance_a:.1f}" if color_distance_a is not None else "N/A"
            dist_b_str = f"{color_distance_b:.1f}" if color_distance_b is not None else "N/A"
            logging.warning(f"DEBUG {tn}:")
            logging.warning(f"  Label-based: sub_a={sub_a}, sub_b={sub_b}, col_a={col_a}, col_b={col_b}")
            logging.warning(f"  Tag bubble: tag_color_a={tag_color_a}, tag_color_b={tag_color_b}")
            logging.warning(f"  Palette fuzzy match:")
            logging.warning(f"    A: {tag_color_a} → {matched_palette_color_a} (dist={dist_a_str}) = {sub_from_color_a}")
            logging.warning(f"    B: {tag_color_b} → {matched_palette_color_b} (dist={dist_b_str}) = {sub_from_color_b}")
            logging.warning(f"  Final subsystems: final_sub_a={final_sub_a}, final_sub_b={final_sub_b}")

        # Treat color changes as subsystem changes (color is part of subsystem definition)
        # Check BOTH label-based colors AND actual tag bubble colors
        sub_changed = final_sub_a != final_sub_b  # Use palette-based subsystems
        label_color_changed = col_a != col_b
        bubble_color_changed = tag_color_a != tag_color_b and tag_color_a is not None and tag_color_b is not None
        
        # More debug logging
        if tn in ["S-98840", "S-98841", "S-98842", "V-98815", "XV-9881521", "XV-9881523"]:
            import logging
            logging.warning(f"  Comparison checks:")
            logging.warning(f"    sub_changed={sub_changed} ({final_sub_a} != {final_sub_b})")
            logging.warning(f"    label_color_changed={label_color_changed} ({col_a} != {col_b})")
            logging.warning(f"    bubble_color_changed={bubble_color_changed} ({tag_color_a} != {tag_color_b})")
            logging.warning(f"    Final: sub_changed={sub_changed} or label_color_changed={label_color_changed} or bubble_color_changed={bubble_color_changed}")
        
        if sub_changed or label_color_changed or bubble_color_changed:
            ct = "subsystem_changed"
        else:
            # Both have no subsystem (None, None) OR exact same subsystem+color
            ct = "unchanged"
        
        if tn in ["S-98840", "S-98841", "S-98842", "V-98815", "XV-9881521", "XV-9881523"]:
            import logging
            logging.warning(f"    Result: ct={ct}")

        if ct == "unchanged":
            unchanged_count += 1
            continue

        action = "review_change"

        rows.append(
            {
                "tag_number": tn,
                "tag_type": ta.get("tag_type") or tb.get("tag_type") or "other",
                "change_type": ct,
                "action_needed": action,
                "subsystem_a": final_sub_a,  # Use palette-based subsystem
                "subsystem_b": final_sub_b,  # Use palette-based subsystem  
                "color_a": col_a or tag_color_a,  # Use label color if available, else tag bubble color
                "color_b": col_b or tag_color_b,
                "tag_bubble_color_a": tag_color_a,  # Store actual bubble colors too
                "tag_bubble_color_b": tag_color_b,
                "drawing_side": "both",
                "page_a": ta.get("page_number"),
                "page_b": tb.get("page_number"),
                "is_x_tag": is_x,
                "drawing_id_a": drawing_a_id,
                "drawing_id_b": drawing_b_id,
                "drawing_number_a": drawing_number_a or "",
                "drawing_number_b": drawing_number_b or "",
                **_view_fields_for_tag(ta, "a"),
                **_view_fields_for_tag(tb, "b"),
            }
        )

    if progress_cb:
        progress_cb(99, "Building summary…")

    subsystem_changed = sum(1 for r in rows if r["change_type"] == "subsystem_changed")
    color_changed = sum(1 for r in rows if r["change_type"] == "color_changed")

    # Debug: count tags with/without subsystem assignments
    tags_with_sub_a = sum(1 for t in tags_a.values() if t.get("nearest_subsystem"))
    tags_with_sub_b = sum(1 for t in tags_b.values() if t.get("nearest_subsystem"))
    tags_with_color_a = sum(1 for t in tags_a.values() if t.get("nearest_subsystem_color"))
    tags_with_color_b = sum(1 for t in tags_b.values() if t.get("nearest_subsystem_color"))
    
    import logging
    logging.warning(f"COMPARISON DEBUG:")
    logging.warning(f"  Tags A: {len(tags_a)}, with subsystem: {tags_with_sub_a}, with color: {tags_with_color_a}")
    logging.warning(f"  Tags B: {len(tags_b)}, with subsystem: {tags_with_sub_b}, with color: {tags_with_color_b}")
    logging.warning(f"  Labels found A: {result_a.get('label_count', 0)}, B: {result_b.get('label_count', 0)}")
    logging.warning(f"  Changes detected: {subsystem_changed}, Unchanged: {unchanged_count}")
    logging.warning(f"  Sample tags A (first 5): {list(tags_a.keys())[:5]}")
    logging.warning(f"  Sample tags B (first 5): {list(tags_b.keys())[:5]}")

    summary = {
        "total_new": sum(1 for r in rows if r["change_type"] == "new"),
        "total_removed": sum(1 for r in rows if r["change_type"] == "removed"),
        "total_unchanged": unchanged_count,
        "total_subsystem_changes": subsystem_changed + color_changed,
        "subsystem_changed": subsystem_changed,
        "color_changed": color_changed,
        "total_tags_a": len(tags_a),
        "total_tags_b": len(tags_b),
        "total_x_tags": sum(1 for r in rows if r.get("is_x_tag")),
        "tags_with_subsystem_a": tags_with_sub_a,
        "tags_with_subsystem_b": tags_with_sub_b,
        "tags_with_color_a": tags_with_color_a,
        "tags_with_color_b": tags_with_color_b,
    }

    debug = {
        "drawing_a": pdf_parser.build_extraction_debug(result_a),
        "drawing_b": pdf_parser.build_extraction_debug(result_b),
        "labels_found_a": result_a.get("label_count", 0),
        "labels_found_b": result_b.get("label_count", 0),
        "subsystems_in_a": result_a.get("subsystems_found") or [],
        "subsystems_in_b": result_b.get("subsystems_found") or [],
    }

    return {
        "comparison_type": "systemized_vs_systemized",
        "drawing_a_id": drawing_a_id,
        "drawing_b_id": drawing_b_id,
        "project_id": project_id,
        "drawing_number_a": drawing_number_a,
        "drawing_number_b": drawing_number_b,
        "tags_only_in_a": sorted(only_a),
        "tags_only_in_b": sorted(only_b),
        "tags_in_both": sorted(both),
        "rows": rows,
        "summary": summary,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "debug": debug,
    }


def compare_drawings(
    drawing_a_path: str,
    drawing_b_path: str,
    role_a: str,
    role_b: str,
    comparison_type: str,
    drawing_a_id: str,
    drawing_b_id: str,
    project_id: str,
    drawing_number_a: str,
    drawing_number_b: str,
    progress_cb: ProgressCb = None,
    tag_exclude_normalized: frozenset[str] | None = None,
) -> dict[str, Any]:
    if comparison_type == "systemized_vs_systemized":
        return compare_systemized_vs_systemized(
            drawing_a_path,
            drawing_b_path,
            drawing_a_id,
            drawing_b_id,
            project_id,
            drawing_number_a,
            drawing_number_b,
            progress_cb=progress_cb,
            tag_exclude_normalized=tag_exclude_normalized,
        )

    if progress_cb:
        progress_cb(5, "Extracting tags from drawing A…")
    tags_a = pdf_parser.extract_tags(
        drawing_a_path,
        tag_exclude_normalized=tag_exclude_normalized,
    )
    if progress_cb:
        progress_cb(22, "Extracting tags from drawing B…")
    tags_b = pdf_parser.extract_tags(
        drawing_b_path,
        tag_exclude_normalized=tag_exclude_normalized,
    )

    sys_a = role_a in ("current_systemized", "prior_systemized")
    sys_b = role_b in ("current_systemized", "prior_systemized")
    if sys_a:
        if progress_cb:
            progress_cb(38, "Resolving colors (drawing A)…")
        tags_a = pdf_parser.extract_colors_near_tags(drawing_a_path, tags_a)
    if sys_b:
        if progress_cb:
            progress_cb(55, "Resolving colors (drawing B)…")
        tags_b = pdf_parser.extract_colors_near_tags(drawing_b_path, tags_b)

    map_a = {t["tag_number"]: t for t in tags_a}
    map_b = {t["tag_number"]: t for t in tags_b}
    set_a, set_b = set(map_a.keys()), set(map_b.keys())

    only_a = set_a - set_b
    only_b = set_b - set_a
    both = set_a & set_b

    rows: list[dict[str, Any]] = []

    def add_row(
        tag_number: str,
        change_type: str,
        side: str,
        ta: dict | None,
        tb: dict | None,
    ) -> None:
        is_x = bool((ta or tb or {}).get("is_x_tag"))
        color_a = (ta or {}).get("color_hex")
        color_b = (tb or {}).get("color_hex")
        prev_sub = None
        action = _action_for_row(change_type, is_x)
        rows.append(
            {
                "tag_number": tag_number,
                "tag_type": (ta or tb or {}).get("tag_type") or "other",
                "change_type": change_type,
                "action_needed": action,
                "subsystem_a": prev_sub,
                "subsystem_b": prev_sub,
                "color_a": color_a,
                "color_b": color_b,
                "drawing_side": side,
                "page_a": (ta or {}).get("page_number"),
                "page_b": (tb or {}).get("page_number"),
                "is_x_tag": is_x,
                "drawing_id_a": drawing_a_id,
                "drawing_id_b": drawing_b_id,
                "drawing_number_a": drawing_number_a or "",
                "drawing_number_b": drawing_number_b or "",
                **_view_fields_for_tag(ta, "a"),
                **_view_fields_for_tag(tb, "b"),
            }
        )

    if progress_cb:
        progress_cb(72, "Building change rows…")

    for tn in sorted(only_a):
        add_row(tn, "removed", "A", map_a.get(tn), None)
    for tn in sorted(only_b):
        add_row(tn, "new", "B", None, map_b.get(tn))
    unchanged_both = 0
    for tn in sorted(both):
        ta, tb = map_a[tn], map_b[tn]
        ct = "unchanged"
        if comparison_type in ("new_vs_systemized", "new_vs_new"):
            ca, cb = _norm_hex(ta.get("color_hex")), _norm_hex(tb.get("color_hex"))
            if ca != cb:
                ct = "color_changed"
        if ct == "unchanged":
            unchanged_both += 1
            continue
        add_row(tn, ct, "both", ta, tb)

    summary = {
        "total_new": sum(1 for r in rows if r["change_type"] == "new"),
        "total_removed": sum(1 for r in rows if r["change_type"] == "removed"),
        "total_unchanged": unchanged_both,
        "total_subsystem_changes": sum(
            1 for r in rows if r["change_type"] in ("subsystem_changed", "color_changed")
        ),
        "total_x_tags": sum(1 for r in rows if r.get("is_x_tag")),
    }

    return {
        "comparison_type": comparison_type,
        "drawing_a_id": drawing_a_id,
        "drawing_b_id": drawing_b_id,
        "project_id": project_id,
        "drawing_number_a": drawing_number_a,
        "drawing_number_b": drawing_number_b,
        "tags_only_in_a": list(only_a),
        "tags_only_in_b": list(only_b),
        "tags_in_both": list(both),
        "rows": rows,
        "summary": summary,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "debug": {
            "drawing_a": {
                "total_tags_extracted": len(tags_a),
                "total_labels_extracted": 0,
                "subsystems_found": [],
                "sample_tags": [],
            },
            "drawing_b": {
                "total_tags_extracted": len(tags_b),
                "total_labels_extracted": 0,
                "subsystems_found": [],
                "sample_tags": [],
            },
        },
    }


def generate_excel_report(comparison_result: dict[str, Any], project_name: str) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Project"
    ws1["B1"] = project_name
    ws1["A2"] = "Comparison type"
    ws1["B2"] = comparison_result.get("comparison_type", "")
    ws1["A3"] = "Date"
    ws1["B3"] = comparison_result.get("generated_at", "")
    s = comparison_result.get("summary") or {}
    ws1["A5"] = "New"
    ws1["B5"] = s.get("total_new", 0)
    ws1["A6"] = "Removed"
    ws1["B6"] = s.get("total_removed", 0)
    ws1["A7"] = "Unchanged (count only; not in sheets below)"
    ws1["B7"] = s.get("total_unchanged", 0)
    ws1["A8"] = "Subsystem/color changes"
    ws1["B8"] = s.get("total_subsystem_changes", 0)
    if "subsystem_changed" in s:
        ws1["A9"] = "Subsystem changed"
        ws1["B9"] = s.get("subsystem_changed", 0)
        ws1["A10"] = "Color changed"
        ws1["B10"] = s.get("color_changed", 0)

    fill_new = PatternFill("solid", fgColor="FFF3CD")
    fill_removed = PatternFill("solid", fgColor="F8D7DA")
    fill_changed = PatternFill("solid", fgColor="CCE5FF")

    ws2 = wb.create_sheet("All Tags")
    headers = [
        "Tag Number",
        "Tag Type",
        "Document (Drawing A)",
        "Document (Drawing B)",
        "Subsystem A",
        "Subsystem B",
        "Color A",
        "Color B",
        "Change Type",
        "Action Needed",
        "Page A",
        "Page B",
        "X-Tag",
    ]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
    dn_a = comparison_result.get("drawing_number_a", "")
    dn_b = comparison_result.get("drawing_number_b", "")
    ncols = len(headers)
    for i, row in enumerate(comparison_result.get("rows") or [], start=2):
        doc_a = row.get("drawing_number_a") or dn_a
        doc_b = row.get("drawing_number_b") or dn_b
        ws2.cell(row=i, column=1, value=row.get("tag_number"))
        ws2.cell(row=i, column=2, value=row.get("tag_type"))
        ws2.cell(row=i, column=3, value=doc_a)
        ws2.cell(row=i, column=4, value=doc_b)
        ws2.cell(row=i, column=5, value=row.get("subsystem_a"))
        ws2.cell(row=i, column=6, value=row.get("subsystem_b"))
        ws2.cell(row=i, column=7, value=row.get("color_a"))
        ws2.cell(row=i, column=8, value=row.get("color_b"))
        ws2.cell(row=i, column=9, value=row.get("change_type"))
        ws2.cell(row=i, column=10, value=row.get("action_needed"))
        ws2.cell(row=i, column=11, value=row.get("page_a"))
        ws2.cell(row=i, column=12, value=row.get("page_b"))
        ws2.cell(row=i, column=13, value="Y" if row.get("is_x_tag") else "")
        ct = row.get("change_type")
        fl = None
        if ct == "new":
            fl = fill_new
        elif ct == "removed":
            fl = fill_removed
        elif ct in ("subsystem_changed", "color_changed"):
            fl = fill_changed
        if fl:
            for col in range(1, ncols + 1):
                ws2.cell(row=i, column=col).fill = fl

    ws3 = wb.create_sheet("Action Required")
    ws3.append(headers)
    for row in comparison_result.get("rows") or []:
        if row.get("action_needed") in (None, "none"):
            continue
        doc_a = row.get("drawing_number_a") or dn_a
        doc_b = row.get("drawing_number_b") or dn_b
        ws3.append(
            [
                row.get("tag_number"),
                row.get("tag_type"),
                doc_a,
                doc_b,
                row.get("subsystem_a"),
                row.get("subsystem_b"),
                row.get("color_a"),
                row.get("color_b"),
                row.get("change_type"),
                row.get("action_needed"),
                row.get("page_a"),
                row.get("page_b"),
                "Y" if row.get("is_x_tag") else "",
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def result_to_json_text(result: dict[str, Any]) -> str:
    return json.dumps(result, default=str)
