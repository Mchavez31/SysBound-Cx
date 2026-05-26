"""
Excel report generation service for tag extraction reports.
Creates Excel files matching PIMS commissioning format.
"""
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_tag_report_excel(
    report_data: List[Dict[str, Any]],
    drawing_number: str,
    output_path: str
) -> str:
    """
    Generate Excel file for tag extraction report.
    
    Args:
        report_data: List of tag dictionaries with keys:
            - plant, module, tag_number, tag_type, tag_description
            - subsystem, discipline, page_number, etc.
        drawing_number: Drawing number for the report
        output_path: Full path where Excel file should be saved
    
    Returns:
        Path to the generated Excel file
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tag Report"
    
    # Define headers matching PIMS format
    headers = [
        'Plant',
        'Module', 
        'Tag No',
        'Tag Type',
        'Tag Description',
        'Disc',  # Discipline
        'Subsystem',
        'P&ID Drawing Name',
        'P&ID Rev',  # NEW: Revision column
        'Page Number',
        'Tag Color',
        'X Position',
        'Y Position'
    ]
    
    # Write headers with styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_num, tag in enumerate(report_data, 2):
        ws.cell(row=row_num, column=1, value=tag.get('plant', 'WOC'))
        ws.cell(row=row_num, column=2, value=tag.get('module', ''))
        ws.cell(row=row_num, column=3, value=tag.get('tag_number', ''))
        ws.cell(row=row_num, column=4, value=tag.get('tag_type', ''))
        ws.cell(row=row_num, column=5, value=tag.get('tag_description', ''))
        ws.cell(row=row_num, column=6, value=tag.get('discipline', ''))
        ws.cell(row=row_num, column=7, value=tag.get('subsystem', ''))
        # Use per-tag drawing number from page extraction (not the parameter)
        ws.cell(row=row_num, column=8, value=tag.get('pid_drawing_number', drawing_number))
        # NEW: Add revision column from per-page extraction
        ws.cell(row=row_num, column=9, value=tag.get('pid_revision', ''))
        ws.cell(row=row_num, column=10, value=tag.get('page_number', ''))
        ws.cell(row=row_num, column=11, value=tag.get('tag_color', ''))
        ws.cell(row=row_num, column=12, value=tag.get('x_position', ''))
        ws.cell(row=row_num, column=13, value=tag.get('y_position', ''))
        
        # Apply borders to data cells
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = thin_border
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = len(headers[col_num - 1])
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            cell = row[0]
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long values
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(['Drawing Number', drawing_number])
    summary_ws.append(['Total Tags', len(report_data)])
    summary_ws.append(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    
    # Count by discipline
    disciplines = {}
    for tag in report_data:
        disc = tag.get('discipline', 'Unknown')
        disciplines[disc] = disciplines.get(disc, 0) + 1
    
    summary_ws.append([])
    summary_ws.append(['Discipline', 'Count'])
    for disc, count in sorted(disciplines.items()):
        summary_ws.append([disc, count])
    
    # Count by tag type
    tag_types = {}
    for tag in report_data:
        tt = tag.get('tag_type', 'Unknown')
        tag_types[tt] = tag_types.get(tt, 0) + 1
    
    summary_ws.append([])
    summary_ws.append(['Tag Type', 'Count'])
    for tt, count in sorted(tag_types.items(), key=lambda x: -x[1])[:20]:  # Top 20
        summary_ws.append([tt, count])
    
    # Save workbook
    wb.save(output_path)
    return output_path
