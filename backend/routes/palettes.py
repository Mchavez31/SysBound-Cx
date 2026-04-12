from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from database_connection import get_db
from models.database import ColorPalette, ProjectMember, User
from routes.auth import get_current_user
import openpyxl
import io

router = APIRouter()

SHEET_MAP = {
    "PFDsPalletColor": "PFD", "PFD Color": "PFD",
    "P&IDs_PalletColor": "P&ID", "Pallet Color P&ID ": "P&ID",
    "SLD_PalletColor": "SLD", "Pallet color SLDs": "SLD",
    "PanelSchedule_PalletColor": "Panel Schedule", "Pallet color Panel Schedule WOC": "Panel Schedule",
    "Telecom_PalletColor": "Telecom", "Pallet color Telecom": "Telecom",
    "AUT_PalletColor ": "Automation", "AUT-BKDs_PalletColor ": "Automation", "Pallet color Automation": "Automation",
    "Infra_PalletColor": "Infrastructure",
}

def extract_palettes_from_xlsx(content: bytes, plant: str) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    entries = []
    for sheet_name in wb.sheetnames:
        dtype = SHEET_MAP.get(sheet_name, sheet_name.strip())
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=120, max_col=8):
            cell_a = row[0]
            if len(row) < 3:
                continue
            fill = cell_a.fill
            hex_color = None
            if fill and fill.fgColor and fill.fgColor.type == "rgb":
                rgb = fill.fgColor.rgb
                if rgb not in ("00000000", "FFFFFFFF", "00FFFFFF", "FF000000"):
                    hex_color = "#" + rgb[2:]
            if hex_color is None:
                continue
            sys_num = row[1].value if len(row) > 1 else None
            if sys_num is None:
                continue
            sys_num = str(sys_num).strip()
            if not sys_num or sys_num in ("System Number", "System \nNumber", "Color"):
                continue
            desc = row[2].value if len(row) > 2 else None
            desc = str(desc).strip() if desc else ""
            r_val = g_val = b_val = None
            for ci in [3, 4, 5]:
                if ci < len(row) and row[ci].value is not None:
                    try:
                        v = int(float(str(row[ci].value)))
                        if 0 <= v <= 255:
                            if r_val is None: r_val = v
                            elif g_val is None: g_val = v
                            elif b_val is None: b_val = v
                    except:
                        pass
            entries.append({
                "plant": plant, "drawing_type": dtype,
                "subsystem_number": sys_num, "subsystem_description": desc,
                "hex_color": hex_color, "r": r_val, "g": g_val, "b": b_val,
            })
    return entries

def check_project_access(project_id: str, user_id: str, db: Session, roles=None):
    q = db.query(ProjectMember).filter(
        ProjectMember.project_id == project_id,
        ProjectMember.user_id == user_id,
    )
    if roles:
        q = q.filter(ProjectMember.role.in_(roles))
    m = q.first()
    if not m:
        raise HTTPException(status_code=403, detail="Not authorized")
    return m

@router.get("/{project_id}")
def list_palettes(project_id: str, plant: Optional[str] = None, drawing_type: Optional[str] = None,
                  db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_project_access(project_id, current_user.id, db)
    q = db.query(ColorPalette).filter(ColorPalette.project_id == project_id)
    if plant:
        q = q.filter(ColorPalette.plant == plant)
    if drawing_type:
        q = q.filter(ColorPalette.drawing_type == drawing_type)
    palettes = q.order_by(ColorPalette.plant, ColorPalette.drawing_type, ColorPalette.subsystem_number).all()
    return [{"id": p.id, "plant": p.plant, "drawing_type": p.drawing_type,
             "subsystem_number": p.subsystem_number, "subsystem_description": p.subsystem_description,
             "hex_color": p.hex_color, "r": p.r, "g": p.g, "b": p.b} for p in palettes]

@router.post("/{project_id}/upload")
async def upload_palette(project_id: str, plant: str, file: UploadFile = File(...),
                         db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # Any project member may upload (same as drawings); viewers included for data imports
    check_project_access(project_id, current_user.id, db)
    content = await file.read()
    entries = extract_palettes_from_xlsx(content, plant)
    if not entries:
        raise HTTPException(status_code=400, detail="No color entries found in file")
    db.query(ColorPalette).filter(ColorPalette.project_id == project_id, ColorPalette.plant == plant).delete()
    for e in entries:
        palette = ColorPalette(project_id=project_id, **e)
        db.add(palette)
    db.commit()
    return {"success": True, "imported": len(entries), "plant": plant}

@router.get("/{project_id}/summary")
def palette_summary(project_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_project_access(project_id, current_user.id, db)
    palettes = db.query(ColorPalette).filter(ColorPalette.project_id == project_id).all()
    summary = {}
    for p in palettes:
        key = f"{p.plant}|{p.drawing_type}"
        if key not in summary:
            summary[key] = {"plant": p.plant, "drawing_type": p.drawing_type, "count": 0}
        summary[key]["count"] += 1
    return list(summary.values())

@router.get("/{project_id}/lookup")
def lookup_color(project_id: str, plant: str, drawing_type: str, subsystem_number: str,
                 db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_project_access(project_id, current_user.id, db)
    p = db.query(ColorPalette).filter(
        ColorPalette.project_id == project_id,
        ColorPalette.plant == plant,
        ColorPalette.drawing_type == drawing_type,
        ColorPalette.subsystem_number == subsystem_number,
    ).first()
    if not p:
        raise HTTPException(status_code=404, detail="Color not found for this combination")
    return {"hex_color": p.hex_color, "r": p.r, "g": p.g, "b": p.b, "subsystem_description": p.subsystem_description}
